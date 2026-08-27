"""
reconciliation/misc_adjustment2_generator.py

MISC Adjustment (2) itemizer for the web app. MISC bills have no INVOICE_NO,
so this keys on the bill reference ("MISC/49/2026"), shared across all four
MISC exports (BILL_REF / BILLNO / BILL_NO).

For each bill:  residual = closing - opening - sales + collection.
A non-zero residual is a MISC adjustment; the residuals sum to MISC Adj(2).
Mirrors the Excel MiscAdjustment2.bas logic exactly.
"""
import csv, io

EPS = 0.005


def _text(src):
    if src is None:
        return None
    if hasattr(src, "read"):
        d = src.read()
        return d.decode("utf-8-sig", "replace") if isinstance(d, (bytes, bytearray)) else d
    if isinstance(src, (bytes, bytearray)):
        return bytes(src).decode("utf-8-sig", "replace")
    with open(src, encoding="utf-8-sig", errors="replace", newline="") as fh:
        return fh.read()


def _rows(src):
    txt = _text(src)
    if txt is None:
        return []
    return list(csv.DictReader(io.StringIO(txt)))


def _num(x):
    x = (x or "").strip().replace(",", "")
    try:
        return float(x) if x else 0.0
    except ValueError:
        return 0.0


def _find(fields, preferred, keys):
    fields = [f for f in (fields or []) if f]
    for f in fields:
        if f.strip().lstrip("\ufeff").upper() == preferred.upper():
            return f
    low = {f.strip().lstrip("\ufeff").lower(): f for f in fields}
    if preferred.lower() in low:
        return low[preferred.lower()]
    for f in fields:
        if any(k in f.strip().lower() for k in keys):
            return f
    return None


def _norm(s):
    return (s or "").strip().upper()


def _classify(ob, cb, sa, co, resid):
    if cb <= EPS and ob > EPS:
        return ("Bill cleared without full collection (write-off / waiver)"
                if co + EPS < ob else "Bill cleared")
    if ob <= EPS and cb > EPS and sa <= EPS:
        return "Bill added without a sale"
    return "Balance reduced (adjustment)" if resid < 0 else "Balance increased (adjustment)"


def detect_misc(open_rows, close_rows, sales_rows, coll_rows):
    open_bal, close_bal, sale_amt, coll_amt = {}, {}, {}, {}
    acc, cat = {}, {}

    def load_debtor(rows, into):
        if not rows:
            return
        h = rows[0].keys()
        cref = _find(h, "BILL_REF", ["bill_ref", "billref"])
        cbal = _find(h, "BALANCE_AMT", ["balance", "outstand"])
        cacc = _find(h, "M_ACCOUNT_NO", ["m_account", "account"])
        ccat = _find(h, "CAT_NAME", ["cat_name", "category"])
        for r in rows:
            k = _norm(r.get(cref))
            if not k:
                continue
            into[k] = into.get(k, 0.0) + _num(r.get(cbal))
            if cacc and k not in acc:
                acc[k] = (r.get(cacc) or "").strip()
            if ccat and k not in cat:
                cat[k] = (r.get(ccat) or "").strip()

    load_debtor(open_rows, open_bal)
    load_debtor(close_rows, close_bal)

    if sales_rows:
        h = sales_rows[0].keys()
        cbill = _find(h, "BILLNO", ["billno", "bill_no"])
        cinv = _find(h, "INVTOT", ["invtot"])
        citm = _find(h, "ITM_AMOUNT", ["itm_amount"])
        cacc = _find(h, "ACNO", ["acno", "account"])
        for r in sales_rows:
            k = _norm(r.get(cbill))
            if not k:
                continue
            if cinv:
                it = _num(r.get(cinv))
                if it > sale_amt.get(k, 0.0):     # one invoice total per bill
                    sale_amt[k] = it
            elif citm:
                sale_amt[k] = sale_amt.get(k, 0.0) + _num(r.get(citm))
            if cacc and k not in acc:
                acc[k] = (r.get(cacc) or "").strip()

    if coll_rows:
        h = coll_rows[0].keys()
        cbill = _find(h, "BILL_NO", ["bill_no", "billno"])
        camt = _find(h, "AMOUNT", ["amount"])
        cacc = _find(h, "M_ACCOUNT_NO", ["m_account", "account"])
        for r in coll_rows:
            k = _norm(r.get(cbill))
            if not k:
                continue
            coll_amt[k] = coll_amt.get(k, 0.0) + _num(r.get(camt))
            if cacc and k not in acc:
                acc[k] = (r.get(cacc) or "").strip()

    keys = set(open_bal) | set(close_bal) | set(sale_amt) | set(coll_amt)
    rows = []
    total = 0.0
    for k in keys:
        ob, cb = open_bal.get(k, 0.0), close_bal.get(k, 0.0)
        sa, co = sale_amt.get(k, 0.0), coll_amt.get(k, 0.0)
        resid = round(cb - ob - sa + co, 2)
        if abs(resid) > EPS:
            rows.append({"bill_ref": k, "account_no": acc.get(k, ""), "category": cat.get(k, ""),
                         "opening": round(ob, 2), "sales": round(sa, 2),
                         "collection": round(co, 2), "closing": round(cb, 2),
                         "adjustment": resid, "reason": _classify(ob, cb, sa, co, resid)})
            total += resid
    rows.sort(key=lambda r: -abs(r["adjustment"]))
    return {"rows": rows, "total": round(total, 2), "n_rows": len(rows)}


def generate_xlsx_bytes(result):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "MISC_Adjustment2"
    hdr = ["Account No", "Bill Ref", "Category", "Opening", "Sales",
           "Collection", "Closing", "Adjustment", "Reason"]
    for c, t in enumerate(hdr, 1):
        ws.cell(1, c, t).font = Font(bold=True)
    r = 2
    for row in result["rows"]:
        ws.cell(r, 1, row["account_no"]); ws.cell(r, 2, row["bill_ref"]); ws.cell(r, 3, row["category"])
        ws.cell(r, 4, row["opening"]); ws.cell(r, 5, row["sales"]); ws.cell(r, 6, row["collection"])
        ws.cell(r, 7, row["closing"]); ws.cell(r, 8, row["adjustment"]); ws.cell(r, 9, row["reason"])
        for c in (4, 5, 6, 7, 8):
            ws.cell(r, c).number_format = "#,##0.00;(#,##0.00)"
        r += 1
    ws.cell(r + 1, 7, "TOTAL Adj(2)").font = Font(bold=True)
    tc = ws.cell(r + 1, 8, result["total"]); tc.font = Font(bold=True); tc.number_format = "#,##0.00;(#,##0.00)"
    for col, w in zip("ABCDEFGHI", (12, 16, 16, 16, 14, 16, 14, 16, 42)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def summary_b64(result):
    import base64, json
    return base64.b64encode(json.dumps(
        {"total": result["total"], "n_rows": result["n_rows"]}).encode()).decode()
