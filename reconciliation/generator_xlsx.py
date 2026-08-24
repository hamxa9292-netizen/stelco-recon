"""
reconciliation/generator_xlsx.py
Excel version of the Debtors Reconciliation Statement.
Consumes calculate(location, figures) -> result (same dict as before) and
produces an .xlsx laid out like the Word statement (electricity + MISC columns).
"""
import io


def _fmt_num(n):
    return round(float(n or 0), 2)


def generate_xlsx(location, result, report_date):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    bold = Font(name="Arial", size=10, bold=True)
    normal = Font(name="Arial", size=10)
    red = Font(name="Arial", size=10, color="CC0000")
    red_bold = Font(name="Arial", size=10, bold=True, color="CC0000")
    grey = PatternFill("solid", fgColor="D9D9D9")
    light = PatternFill("solid", fgColor="EEEEEE")
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")
    thin_top = Border(top=Side(style="thin"))
    dbl_bottom = Border(bottom=Side(style="double"))

    # Header block
    ws["A1"] = "Customer Services and Billing Department"; ws["A1"].font = bold
    ws["A2"] = "    STELCO"; ws["A2"].font = normal
    ws["A4"] = f"Date: {report_date}"; ws["A4"].font = normal
    ws["A6"] = f"Debtors Reconciliation Statement - {result['location_name']}"
    ws["A6"].font = Font(name="Arial", size=11, bold=True, underline="single")
    ws["A7"] = report_date; ws["A7"].font = bold

    # Column headers
    hr = 9
    ws.cell(hr, 2, "ELECTRICITY (MRF)").font = bold
    ws.cell(hr, 3, "MISC. (MRF)").font = bold
    ws.cell(hr, 2).alignment = center
    ws.cell(hr, 3).alignment = center
    ws.cell(hr, 2).fill = grey
    ws.cell(hr, 3).fill = grey
    ws.cell(hr, 1).fill = grey

    r = hr + 1
    for row in result["rows"]:
        is_sub = row.get("subtotal", False)
        is_final = row.get("final", False)
        is_bold = row.get("bold", False) or is_sub or is_final
        ws.cell(r, 1, row["label"]).font = bold if is_bold else normal
        for ci, key in ((2, "elec"), (3, "misc")):
            v = _fmt_num(row[key])
            cell = ws.cell(r, ci, v)
            cell.number_format = "#,##0.00;(#,##0.00)"
            cell.alignment = right
            if v < 0:
                cell.font = red_bold if is_bold else red
            else:
                cell.font = bold if is_bold else normal
            if is_sub:
                cell.fill = light; cell.border = thin_top
            if is_final:
                cell.fill = grey; cell.border = dbl_bottom
        if is_sub:
            ws.cell(r, 1).fill = light
        if is_final:
            ws.cell(r, 1).fill = grey
        r += 1

    # footnote
    r += 1
    ws.cell(r, 1, "*Credit invoice in the Total Sales").font = Font(name="Arial", size=8, italic=True, color="555555")

    # signatures
    r += 3
    sigs = [("Prepared By:", "Hamza Abdul Sattar", "Admin. Supervisor"),
            ("Checked By:", "Ali Amir", "Deputy Service Manager"),
            ("Approved By:", "Hussain Waheed", "General Manager")]
    for ci, (lab, name, title) in enumerate(sigs, start=1):
        ws.cell(r, ci, lab).font = bold
        ws.cell(r + 4, ci, name).font = bold
        ws.cell(r + 5, ci, title).font = normal

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
