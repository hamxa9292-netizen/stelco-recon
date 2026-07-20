"""
reconciliation/adjustment2_generator.py   (v3 - timing/over-payment detector)

Itemises Adjustment (2) into invoice-level rows via three deterministic classes,
then reconciles their sum to the identity plug.

    Adj(2) = Closing - Opening(after adj) - Sales - Credits + Collection   (= the plug)

Classes:
  A  late payment on a current-period sale
       closing debtor row with BILL_DATE after month-end AND PAY_AMT > 0
       amount = -PAY_AMT        (exact, from the closing debtor file)
  C  over-payment / duplicate on a settled past bill
       realised ORD=1 payment on a past/current-period invoice that is absent
       from BOTH debtor ledgers and NOT billed in the current sales report
       amount = +payment        (exact, from the collection file)
  B  early payment on a future-dated bill
       realised ORD=1 payment whose BILL_REF period is AFTER the recon month
       amount = the future bill value -> NOT in current files, so it is
       BACKED OUT of the plug:  Class_B_total = plug - sum(A) - sum(C)
       exact when there is a single Class B invoice; when there are 2+, the
       combined total is known but the per-invoice split needs next-month sales
       (supply next_sales_amounts) or is left for manual entry.

Inputs (paths, text streams, or bytes): opening debtor, closing debtor,
current collection, current sales, and the four control totals for the plug.

Requires: openpyxl==3.1.5
"""

import csv, io, json, base64, calendar, datetime
from dataclasses import dataclass, field
from typing import Optional, List, Tuple, Dict

# column names (auto-detected if absent)
INV="INVOICE_NO"; BAL="BALANCE_AMT"; BILLDATE="BILL_DATE"; PAYAMT="PAY_AMT"
CAMT="COLLECT_AMOUNT"; CREF="BILL_REF"; CORD="ORD"; CCANCEL="CANCEL_DATE"
CACC="ACCOUNT_NO"; SAMT="AMOUNT"
EPS=0.005

def _num(x):
    x=(x or "").strip().replace(",","")
    try: return float(x) if x else 0.0
    except ValueError: return 0.0

def _text(src):
    if hasattr(src,"read"):
        d=src.read(); return d.decode("utf-8-sig","replace") if isinstance(d,(bytes,bytearray)) else (d.lstrip("\ufeff") if isinstance(d,str) else d)
    if isinstance(src,(bytes,bytearray)): return bytes(src).decode("utf-8-sig","replace")
    with open(src, encoding="utf-8-sig", errors="replace", newline="") as f: return f.read()

def _rows(src): return list(csv.DictReader(io.StringIO(_text(src))))

def _find(fields, preferred, keys):
    fields=[f for f in (fields or []) if f]
    for f in fields:
        if f.strip()==preferred: return f
    low={f.strip().lower():f for f in fields}
    if preferred.lower() in low: return low[preferred.lower()]
    for f in fields:
        if any(k in f.strip().lower() for k in keys): return f
    return None

def _date(s):
    s=(s or "").strip()
    for fmt in ("%d-%m-%Y","%Y-%m-%d","%d/%m/%Y","%m/%d/%Y"):
        try: return datetime.datetime.strptime(s,fmt).date()
        except ValueError: pass
    return None

def _period(ref):
    ref=(ref or "").strip()
    if "/" not in ref: return None
    try:
        y,rest=ref.split("/",1); return (int(y), int(rest.split("-")[0]))
    except (ValueError,IndexError): return None

def _month_end(y,m):
    return datetime.date(y,m,calendar.monthrange(y,m)[1])

# ---------------------------------------------------------------- loaders
def load_collection_rows(src): return _rows(src)

def load_debtor_rows(src):
    return _rows(src)

def debtor_invoice_set(rows, positive_only=True):
    if not rows: return set()
    inv=_find(rows[0].keys(), INV, ["invoice"]); bal=_find(rows[0].keys(), BAL, ["balance","outstand"])
    out=set()
    for r in rows:
        i=(r.get(inv) or "").strip()
        if not i: continue
        if bal is None: out.add(i)
        else:
            b=_num(r.get(bal))
            if (b>EPS) if positive_only else (abs(b)>EPS): out.add(i)
    return out

def sales_invoice_set(rows):
    if not rows: return set()
    inv=_find(rows[0].keys(), INV, ["invoice"])
    return {(r.get(inv) or "").strip() for r in rows if (r.get(inv) or "").strip()}

def sales_amounts(rows):
    if not rows: return {}
    inv=_find(rows[0].keys(), INV, ["invoice"]); amt=_find(rows[0].keys(), SAMT, ["amount"])
    out={}
    for r in rows:
        i=(r.get(inv) or "").strip()
        if i: out[i]=out.get(i,0.0)+_num(r.get(amt))
    return out

# ---------------------------------------------------------------- core
@dataclass
class Adj2Result:
    recon_month: Tuple[int,int]
    plug: Optional[float]
    rows: List[dict] = field(default_factory=list)
    total: float = 0.0
    residual: Optional[float] = None
    classB_needs_split: bool = False
    note: str = ""

def detect(coll_rows, open_rows, close_rows, sales_rows, recon_month,
           plug=None, next_sales_amounts: Optional[Dict[str,float]]=None):
    y,m=recon_month
    month_end=_month_end(y,m)
    d_open=debtor_invoice_set(open_rows)
    d_close=debtor_invoice_set(close_rows)
    billed=sales_invoice_set(sales_rows)

    # column handles for closing debtor (Class A)
    cf=close_rows[0].keys() if close_rows else []
    c_inv=_find(cf,INV,["invoice"]); c_bd=_find(cf,BILLDATE,["bill_date","billdate"])
    c_pay=_find(cf,PAYAMT,["pay_amt","payamt","paid"]); c_acc=_find(cf,CACC,["account"]); c_ref=_find(cf,CREF,["bill_ref","ref"])

    # per-invoice realised collection in the CURRENT month (ORD=1, cancelled excluded)
    coll_by_inv={}; coll_meta={}
    for r in coll_rows:
        if (r.get(CORD) or "").strip()=="1" and not (r.get(CCANCEL) or "").strip():
            iv=(r.get(INV) or "").strip()
            if iv:
                coll_by_inv[iv]=coll_by_inv.get(iv,0.0)+_num(r.get(CAMT))
                coll_meta.setdefault(iv, r)

    rows=[]
    # CLASS A: bill dated after month-end with a payment already applied, where that
    # payment is NOT (fully) in this month's collection -> a genuine timing gap.
    # If the payment IS in this month's collection, the invoice reconciles normally
    # (sale, payment and closing balance all in-period) and is NOT an adjustment.
    if c_bd and c_pay:
        for r in close_rows:
            bd=_date(r.get(c_bd)); pay=_num(r.get(c_pay))
            if bd and bd>month_end and pay>EPS:
                iv=(r.get(c_inv) or "").strip()
                gap=round(pay - coll_by_inv.get(iv,0.0), 2)   # payment not covered by this month's collection
                if gap>EPS:
                    rows.append({"invoice_no":iv,
                                 "account_no":(r.get(c_acc) or "").strip() if c_acc else "",
                                 "bill_ref":(r.get(c_ref) or "").strip() if c_ref else "",
                                 "amount":round(-gap,2), "class":"A",
                                 "reason":"late payment (paid, not in this month's collection)"})
    # CLASS B & CLASS C from collection
    B=[]; Ctot={}
    for r in coll_rows:
        if (r.get(CORD) or "").strip()!="1" or (r.get(CCANCEL) or "").strip(): continue
        inv=(r.get(INV) or "").strip()
        if not inv: continue
        p=_period(r.get(CREF))
        if p and p>recon_month:
            B.append({"invoice_no":inv,"account_no":(r.get(CACC) or "").strip(),
                      "bill_ref":(r.get(CREF) or "").strip()})
        elif p and p<=recon_month and inv not in d_open and inv not in d_close and inv not in billed:
            key=(inv,(r.get(CACC) or "").strip(),(r.get(CREF) or "").strip())
            Ctot[key]=Ctot.get(key,0.0)+_num(r.get(CAMT))
    for (inv,acc,ref),a in Ctot.items():
        rows.append({"invoice_no":inv,"account_no":acc,"bill_ref":ref,
                     "amount":round(a,2),"class":"C","reason":"over-payment on settled bill"})

    # dedupe B by invoice
    seen=set(); Bu=[]
    for b in B:
        if b["invoice_no"] not in seen:
            seen.add(b["invoice_no"]); Bu.append(b)

    sumAC=round(sum(r["amount"] for r in rows),2)
    needs_split=False; note=""
    if Bu:
        if plug is not None:
            B_total=round(plug - sumAC,2)
            if len(Bu)==1:
                Bu[0].update({"amount":B_total,"class":"B","reason":"early payment (future-dated bill)"})
                rows.append(Bu[0])
            else:
                # split per-invoice from next-month sales if available; else flag
                assigned=0.0; unresolved=[]
                for b in Bu:
                    amt=(next_sales_amounts or {}).get(b["invoice_no"])
                    if amt is not None:
                        b.update({"amount":round(amt,2),"class":"B","reason":"early payment (future-dated bill)"}); assigned+=amt; rows.append(b)
                    else:
                        unresolved.append(b)
                gap=round(B_total-round(assigned,2),2)
                if unresolved:
                    needs_split=True
                    note=(f"{len(unresolved)} Class B invoice(s) need next-month sales to split "
                          f"{gap:.2f}: "+", ".join(b['invoice_no'] for b in unresolved))
                    for b in unresolved:
                        b.update({"amount":None,"class":"B","reason":"early payment (future-dated bill) - amount pending next-month sales"}); rows.append(b)
        else:
            for b in Bu:
                b.update({"amount":None,"class":"B","reason":"early payment (future-dated bill) - amount pending (no plug supplied)"}); rows.append(b)
            needs_split=True; note="No plug supplied; Class B amounts unresolved."

    # CLASS D - over-payment on a CURRENT-month bill, resolved by exact plug match.
    #
    # Many invoices are over-paid by small amounts each month (customers paying a
    # round figure). The billing system books most of those excesses to Credits -
    # a separate line on the statement - so they are NOT Adj(2). The decisive
    # field is the applied-vs-credit split on the payment, which no available
    # export carries. The excess alone therefore cannot identify the adjustment:
    # a threshold that catches one month's row wrongly catches another's.
    #
    # What CAN be trusted is an exact match. If precisely ONE candidate's excess
    # equals the outstanding residual to the cent, that is the adjustment. Zero
    # or several matches are reported as candidates for review, never guessed.
    d_candidates=[]
    if plug is not None:
        so_far=round(sum(r["amount"] for r in rows if r["amount"] is not None),2)
        gap_left=round(plug-so_far,2)
        if abs(gap_left)>EPS:
            sale_amt=sales_amounts(sales_rows)
            for iv,s_amt in sale_amt.items():
                if s_amt<=EPS or iv in d_open or iv in d_close:
                    continue
                excess=round(coll_by_inv.get(iv,0.0)-s_amt,2)
                if excess>EPS:
                    d_candidates.append((iv,excess))
            exact=[c for c in d_candidates if abs(c[1]-gap_left)<=EPS]
            if len(exact)==1:
                iv,amt=exact[0]
                mrow=coll_meta.get(iv,{})
                rows.append({"invoice_no":iv,
                             "account_no":(mrow.get(CACC) or "").strip(),
                             "bill_ref":(mrow.get(CREF) or "").strip(),
                             "amount":amt,"class":"D",
                             "reason":"over-payment on current-month bill"})
            elif d_candidates:
                d_candidates.sort(key=lambda c:-c[1])
                top=", ".join(f"{iv} ({ex:.2f})" for iv,ex in d_candidates[:5])
                if len(exact)>1:
                    note=(note+" " if note else "")+(
                        f"{len(exact)} current-month over-payments each match the "
                        f"{gap_left:.2f} residual exactly - ambiguous, review needed: "
                        +", ".join(iv for iv,_ in exact))
                    needs_split=True
                else:
                    note=(note+" " if note else "")+(
                        f"Residual {gap_left:.2f} unmatched. Nearest current-month "
                        f"over-payments: {top}. The applied-vs-credit split is not "
                        f"exported, so this needs manual confirmation.")
                    needs_split=True

    total=round(sum(r["amount"] for r in rows if r["amount"] is not None),2)
    residual=round(plug-total,2) if plug is not None else None
    rows.sort(key=lambda r:(r["class"], -(abs(r["amount"]) if r["amount"] is not None else 0)))
    return Adj2Result(recon_month, plug, rows, total, residual, needs_split, note)

def summary_b64(res: Adj2Result):
    steps=[
        {"title":"Class A - late payments","desc":"paid after sale period","val":f"{sum(1 for r in res.rows if r['class']=='A')} row(s)"},
        {"title":"Class C - over-payments","desc":"settled past bill","val":f"{sum(1 for r in res.rows if r['class']=='C')} row(s)"},
        {"title":"Class B - future bills","desc":"paid before sale; backed out of plug","val":f"{sum(1 for r in res.rows if r['class']=='B')} row(s)"},
        {"title":"Reconcile to plug","desc":"residual should be 0","val":f"total {res.total:.2f} / plug {('%.2f'%res.plug) if res.plug is not None else '—'}"},
    ]
    payload={"recon_month":f"{res.recon_month[0]}-{res.recon_month[1]:02d}","plug":res.plug,
             "total":res.total,"residual":res.residual,"row_count":len(res.rows),
             "needs_split":res.classB_needs_split,"note":res.note,"steps":steps}
    return base64.b64encode(json.dumps(payload).encode()).decode()

def generate_xlsx_bytes(res: Adj2Result):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb=Workbook(); ws=wb.active; ws.title="Adjustment2"; b=Font(bold=True)
    ws.append(["Invoice No","Account No","Bill Ref","Amount","Class","Reason"])
    for c in ws[1]: c.font=b
    for r in res.rows:
        ws.append([r["invoice_no"],r["account_no"],r["bill_ref"],r["amount"],r["class"],r["reason"]])
    ws.append([]); ws.append(["TOTAL","","",res.total,"",""]); ws[ws.max_row][0].font=b
    ws.append(["PLUG (identity)","","",res.plug,"",""])
    ws.append(["RESIDUAL","","",res.residual,"",""])
    if res.note: ws.append(["NOTE","","",res.note,"",""])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


# ---------------------------------------------------------------- plug helpers
def realised_collection(coll_rows):
    """Realised collection = ORD 1+2 payments, cancelled excluded."""
    t = 0.0
    for r in coll_rows:
        if (r.get(CORD) or "").strip() in ("1", "2") and not (r.get(CCANCEL) or "").strip():
            t += _num(r.get(CAMT))
    return round(t, 2)


def plug_from_totals(opening_after_adj, closing, sales, credits, collection):
    """Adj(2) plug = Closing - Opening(after adj) - Sales - Credits + Collection."""
    return round(closing - opening_after_adj - sales - credits + collection, 2)
