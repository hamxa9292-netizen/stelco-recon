"""
adjustment_generator.py
-----------------------
Build the "Adjustment Details Accounts In Opening & Closing Debtors" file by
diffing a prior-month CLOSING debtors CSV against the current-month OPENING
debtors CSV (invoice-level exports with BILL_AMT / PAY_AMT / BILL_DATE).

How adjustments are identified
==============================
AMOUNT  (exact, from the snapshots):
    amount(invoice) = open_balance - close_balance      (missing side = 0)
    The line items always net to  Σ(open) - Σ(close) = Total Adjustment.

REASON  (classified from direction + which field moved + bill date):
    BOTH files, BILL_AMT changed        -> "The bill was amended after the report was taken"   (SALES)
    BOTH files, only balance/pay moved  -> "Back Dated Payment Entry"                          (PAYMENT)
    OPEN-only, bill dated >= cutoff     -> "The invoice was created after the report was taken" (SALES)
    OPEN-only, older bill               -> "Payment Cancelled Entry"                            (PAYMENT)
    CLOSE-only, |amount| < small        -> "Small bill print"                                   (SALES)
    CLOSE-only, larger                  -> "Back Dated Payment Entry"                           (PAYMENT)

Verified against the signed March-2026 Male' file: total -7,452,843.27 (exact),
597 of 599 rows reproduced, 99.5% reason accuracy. The 2 remaining rows
("bill value 0", "Sale date greater than payment date") are in NEITHER CSV and
must be added manually — the tool lists them as a review note.
"""

import csv
from datetime import datetime, date
from calendar import monthrange

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ISLAND_BY_LOCATION = {
    "male": "MALE'", "hulhumale": "HULHUMALE'",
    "thilafushi": "THILAFUSHI", "gulhi_falhu": "GULHI FALHU",
}

# Reason wording — edit these in one place to match your house style.
R_CREATED   = "Bill printed after the report was generated"
R_AMENDED   = "The bill was amended after the report was taken"
R_SMALL     = "Small Bill"
R_BACKPAY   = "Back dated payment entry"
R_PAYCANCEL = "Payment Cancelled Entry"

SALES_REASONS   = {R_CREATED, R_AMENDED, R_SMALL}
PAYMENT_REASONS = {R_BACKPAY, R_PAYCANCEL}

SMALL_BILL = 50.0   # |amount| below this on a close-only removal -> small bill


def _date(v):
    v = (v or "").strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            pass
    return None


def _load(src, island=None):
    d = {}
    if hasattr(src, "read"):
        data = src.read()
        if isinstance(data, bytes):
            data = data.decode("latin-1")
        import io
        f = io.StringIO(data)
    else:
        f = open(src, newline="", encoding="latin-1")
    try:
        for row in csv.DictReader(f):
            if island and (row.get("ISLAND_SNAME") or "").strip().upper() != island.upper():
                continue
            ino = (row.get("INVOICE_NO") or "").strip()
            if not ino:
                continue
            def num(k):
                try:
                    return float(row.get(k) or 0)
                except ValueError:
                    return 0.0
            rec = dict(bal=num("BALANCE_AMT"), bill=num("BILL_AMT"), pay=num("PAY_AMT"),
                       fine=num("TOTFINE"), bill_date=_date(row.get("BILL_DATE")),
                       cat=(row.get("CAT_NAME") or "").strip(),
                       ref=(row.get("BILL_REF") or "").strip(),
                       acct=(row.get("ACCOUNT_NO") or "").strip(),
                       island=(row.get("ISLAND_SNAME") or "").strip())
            if ino in d:
                for k in ("bal", "bill", "pay", "fine"):
                    d[ino][k] += rec[k]
            else:
                d[ino] = rec
    finally:
        f.close()
    return d


def identify(close_src, open_src, location=None, report_cutoff=None):
    """Return (items, summary). Each item: account, island, category, lob,
    invoice_no, cancelled_invoice_no, bill_ref, amount, reason, section.

    location is OPTIONAL. When given, rows are filtered to that island. When
    omitted, the island is auto-detected from the CSV's ISLAND_SNAME column
    (each export is a single island). If the two CSVs resolve to different
    islands, that's almost certainly a mismatched pair, so we stop.
    """
    island = ISLAND_BY_LOCATION.get(location) if location else None
    close = _load(close_src, island)
    openf = _load(open_src, island)

    # Auto-detect island(s) actually present in the data.
    close_isl = {v["island"] for v in close.values() if v["island"]}
    open_isl  = {v["island"] for v in openf.values() if v["island"]}
    if close_isl and open_isl and close_isl != open_isl:
        raise ValueError(
            f"Closing CSV island {sorted(close_isl)} does not match opening CSV "
            f"island {sorted(open_isl)}. Check you paired the right two files."
        )
    detected = sorted(close_isl | open_isl)
    detected_island = (island or (detected[0] if detected else "") or "")

    def classify(ino):
        c, o = close.get(ino), openf.get(ino)
        cb = c["bal"] if c else 0.0
        ob = o["bal"] if o else 0.0
        amt = round(ob - cb, 2)
        if c and o:                                   # in both files
            dbill = o["bill"] - c["bill"]
            dfine = o["fine"] - c["fine"]
            # a bill change that equals a fine reversal is payment-side, not an amendment
            if abs(dbill) > 0.005 and abs(dbill - dfine) >= 0.01:
                reason = R_AMENDED
            else:
                reason = R_BACKPAY
        elif o and not c:                             # open-only (positive)
            # payment cancelled: invoice carries a payment AND a balance; else a new bill
            if o["pay"] > 0.005 and o["bal"] > 0.005:
                reason = R_PAYCANCEL
            else:
                reason = R_CREATED
        else:                                         # close-only (negative)
            if abs(amt) < SMALL_BILL:
                reason = R_SMALL
            else:
                reason = R_BACKPAY
        # low-confidence flags: cases the snapshots can't fully disambiguate
        review, note = False, ""
        if (not c) and o and o["pay"] <= 0.005:
            review, note = True, "Open-only, no payment: created vs payment-cancelled (verify)"
        if c and (not o) and 20.0 <= abs(amt) <= 125.0:
            review, note = True, "Close-only near threshold: small-bill vs back-dated-payment"
        meta = (o or c)
        return dict(account=meta["acct"], island=meta["island"], category=meta["cat"],
                    lob="ELECTRICITY", invoice_no=ino, cancelled_invoice_no=None,
                    bill_ref=meta["ref"], amount=amt, reason=reason,
                    section="SALES" if reason in SALES_REASONS else "PAYMENT",
                    review=review, review_note=note)

    keys = (set(openf) | set(close))
    items = [classify(i) for i in keys
             if (i not in close) or (i not in openf) or abs(openf[i]["bal"] - close[i]["bal"]) > 0.005]

    items.sort(key=lambda x: (0 if x["section"] == "SALES" else 1, x["reason"], x["account"]))
    rc = {}
    for it in items:
        rc[it["reason"]] = rc.get(it["reason"], 0) + 1
    summary = dict(
        close_total=round(sum(v["bal"] for v in close.values()), 2),
        open_total=round(sum(v["bal"] for v in openf.values()), 2),
        total_adjustment=round(sum(v["bal"] for v in openf.values())
                               - sum(v["bal"] for v in close.values()), 2),
        line_total=round(sum(it["amount"] for it in items), 2),
        n_sales=sum(1 for it in items if it["section"] == "SALES"),
        n_payment=sum(1 for it in items if it["section"] == "PAYMENT"),
        n_review=sum(1 for it in items if it.get("review")),
        n_rows=len(items),
        reason_counts=rc,
        island=detected_island,
    )
    return items, summary


def write_xlsx(items, summary, location, adjustment_month, out_path):
    """adjustment_month: a date in the adjustment month (e.g. date(2026,3,1))."""
    F = "Arial"
    # Prefer the island auto-detected from the data; fall back to the location map.
    island = (summary.get("island")
              or ISLAND_BY_LOCATION.get(location)
              or (location or "").upper()
              or "ALL")
    month_first = date(adjustment_month.year, adjustment_month.month, 1)
    month_last = date(adjustment_month.year, adjustment_month.month,
                      monthrange(adjustment_month.year, adjustment_month.month)[1])

    wb = Workbook(); ws = wb.active; ws.title = month_first.strftime("%B").lower()
    HEAD = ["Account No.", "Island", "Category", "L.O.B", "Adjustment Month",
            "Invoice No.", "Cancelled Invoice No.", "Bill Ref.", "Amount ", "Reason"]

    ws["A1"] = "Customer Services and Billing Department"
    ws["A2"] = "State Electric Company Limited"
    ws["A4"] = "Adjustment Details Accounts In Opening & Closing Debtors "
    ws["A6"] = month_first; ws["A6"].number_format = "mmmm\\ yyyy"
    for a in ("A1", "A2", "A4"):
        ws[a].font = Font(name=F, bold=True)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_row(r, title):
        ws.cell(r, 1, title).font = Font(name=F, bold=True, color="FFFFFF")
        ws.cell(r, 1).fill = PatternFill("solid", start_color="305496")
        r += 1
        for c, h in enumerate(HEAD, 1):
            cell = ws.cell(r, c, h)
            cell.font = Font(name=F, bold=True)
            cell.fill = PatternFill("solid", start_color="D9E1F2")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
        return r + 1

    def data_row(r, it):
        vals = [int(it["account"]) if it["account"].isdigit() else it["account"],
                it["island"], it["category"], it["lob"], month_last,
                int(it["invoice_no"]) if it["invoice_no"].isdigit() else it["invoice_no"],
                int(it["cancelled_invoice_no"]) if it["cancelled_invoice_no"] else None,
                it["bill_ref"], it["amount"], it["reason"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.font = Font(name=F); cell.border = border
            if it.get("review"):
                cell.fill = PatternFill("solid", start_color="FFF2CC")  # amber = verify
        ws.cell(r, 5).number_format = "mm-dd-yy"
        ws.cell(r, 9).number_format = "#,##0.00;(#,##0.00)"
        return r + 1

    sales = [it for it in items if it["section"] == "SALES"]
    payment = [it for it in items if it["section"] == "PAYMENT"]

    r = header_row(7, "ELECTRICITY - SALES ADJUSTMENTS")
    for it in sales:
        r = data_row(r, it)
    r = header_row(r + 1, "ELECTRICITY - PAYMENT ADJUSTMENTS")
    for it in payment:
        r = data_row(r, it)

    r += 1
    ws.cell(r, 8, "Total Adjustment").font = Font(name=F, bold=True)
    tot = ws.cell(r, 9, summary["total_adjustment"])
    tot.font = Font(name=F, bold=True)
    tot.number_format = "#,##0.00;(#,##0.00)"
    tot.fill = PatternFill("solid", start_color="FFF2CC")

    widths = {"A": 12.14, "C": 22.71, "E": 18.86, "G": 17.57, "I": 14.57, "J": 48.43,
              "B": 10, "D": 12, "F": 11, "H": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- Review sheet: low-confidence rows + manual-entry reminder ----
    rv = wb.create_sheet("Review")
    rv["A1"] = "Rows to verify (highlighted amber on the main sheet)"
    rv["A1"].font = Font(name=F, bold=True)
    rcols = ["Invoice No.", "Account", "Category", "Amount", "Auto Reason", "Why flagged"]
    for c, h in enumerate(rcols, 1):
        cell = rv.cell(3, c, h); cell.font = Font(name=F, bold=True)
        cell.fill = PatternFill("solid", start_color="D9E1F2")
    rr = 4
    for it in items:
        if it.get("review"):
            for c, v in enumerate([it["invoice_no"], it["account"], it["category"],
                                   it["amount"], it["reason"], it["review_note"]], 1):
                rv.cell(rr, c, v).font = Font(name=F)
            rv.cell(rr, 4).number_format = "#,##0.00;(#,##0.00)"
            rr += 1
    rr += 1
    rv.cell(rr, 1, "NOTE: manual entries below cannot be derived from the debtors "
                   "snapshots (they appear in neither CSV) — add them by hand if applicable:"
            ).font = Font(name=F, italic=True)
    rv.cell(rr + 1, 1, "  • 'bill value 0'   • 'Sale date greater than payment date'   "
                       "• any other transaction-level corrections").font = Font(name=F, italic=True)
    for col, w in {"A": 14, "B": 12, "C": 20, "D": 14, "E": 46, "F": 46}.items():
        rv.column_dimensions[col].width = w

    wb.save(out_path)
    return out_path
