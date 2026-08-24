from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
import tempfile, os, json, io
from parser.male import parse_male
from parser.hulhumale import parse_hulhumale
from parser.thilafushi import parse_thilafushi
from parser.gulhi_falhu import parse_gulhi_falhu
from parser.adjustments import find_adjustments, ISLAND_BY_LOCATION
from reconciliation.adjustment_generator import identify, write_xlsx
from reconciliation.adjustment2_generator import (
    load_collection_rows, load_debtor_rows, detect,
    realised_collection, plug_from_totals,
    generate_xlsx_bytes, summary_b64,
)
from datetime import datetime, date
from reconciliation.calculator import calculate
from reconciliation.generator import generate_docx
from reconciliation.csv_parser import parse_csv_figures
from reconciliation.generator_xlsx import generate_xlsx
app = FastAPI(title="STELCO Debtors Reconciliation API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Adjustment-Total", "X-Adjustment-Rows",
                    "X-Adjustment-Review", "X-Adjustment-Summary",
                    "X-Adjustment2-Summary"],
)
PARSERS = {
    "male":       parse_male,
    "hulhumale":  parse_hulhumale,
    "thilafushi": parse_thilafushi,
    "gulhi_falhu":parse_gulhi_falhu,
    "other_islands": parse_thilafushi,   # same reports/formula as Thilafushi
}
@app.get("/")
def root():
    return {"status": "STELCO Recon API running"}
@app.post("/parse")
async def parse_files(
    location: str = Form(...),
    open_pdf:       UploadFile = File(...),
    close_pdf:      UploadFile = File(...),
    sales_pdf:      UploadFile = File(...),
    misc_open_pdf:  UploadFile = File(None),
    misc_close_pdf: UploadFile = File(None),
    misc_sales_pdf: UploadFile = File(None),
    recon_pdf:      UploadFile = File(...),
    collection_pdf: UploadFile = File(...),
    cash_collection_pdf: UploadFile = File(None),   # Hulhumale' only
    billing_pdf:    UploadFile = File(None),        # Hulhumale' only
):
    """
    Step 1 — Parse all uploaded PDFs and return extracted figures for review.
    """
    if location not in PARSERS:
        raise HTTPException(400, f"Unknown location: {location}")
    # Save uploads to temp files
    files = {}
    uploads = {
        "open":            open_pdf,
        "close":           close_pdf,
        "sales":           sales_pdf,
        "misc_open":       misc_open_pdf,
        "misc_close":      misc_close_pdf,
        "misc_sales":      misc_sales_pdf,
        "recon":           recon_pdf,
        "collection":      collection_pdf,
        "cash_collection": cash_collection_pdf,
        "billing":         billing_pdf,
    }
    tmp_dir = tempfile.mkdtemp()
    for key, upload in uploads.items():
        if upload is None:
            files[key] = None
            continue
        path = os.path.join(tmp_dir, f"{key}.pdf")
        content = await upload.read()
        with open(path, "wb") as f:
            f.write(content)
        files[key] = path
    try:
        figures = PARSERS[location](files)
        return {"location": location, "figures": figures}
    except Exception as e:
        raise HTTPException(500, f"Parse error: {str(e)}")
@app.post("/parse_csv")
async def parse_csv(
    location: str = Form(...),
    open_csv:        UploadFile = File(...),
    close_csv:       UploadFile = File(...),
    sales_csv:       UploadFile = File(...),
    collection_csv:  UploadFile = File(...),
    credits_csv:     UploadFile = File(None),
    prior_close_csv: UploadFile = File(None),
    misc_open_csv:   UploadFile = File(None),
    misc_close_csv:  UploadFile = File(None),
    misc_sales_csv:  UploadFile = File(None),
    misc_coll_csv:   UploadFile = File(None),
    blueridge: float = Form(None),   # Hulhumale' — typed from PDF
    wamco:     float = Form(None),
    elec_bf:   float = Form(None),   # prior signed b/f, if no prior_close_csv
):
    """CSV replacement for /parse — returns the same `figures` dict for review."""
    if location not in PARSERS:
        raise HTTPException(400, f"Unknown location: {location}")
    try:
        files = {
            "open_csv":        await open_csv.read(),
            "close_csv":       await close_csv.read(),
            "sales_csv":       await sales_csv.read(),
            "collection_csv":  await collection_csv.read(),
            "credits_csv":     await credits_csv.read() if credits_csv else None,
            "prior_close_csv": await prior_close_csv.read() if prior_close_csv else None,
            "misc_open_csv":   await misc_open_csv.read() if misc_open_csv else None,
            "misc_close_csv":  await misc_close_csv.read() if misc_close_csv else None,
            "misc_sales_csv":  await misc_sales_csv.read() if misc_sales_csv else None,
            "misc_coll_csv":   await misc_coll_csv.read() if misc_coll_csv else None,
        }
        passthrough = {"blueridge": blueridge, "wamco": wamco, "elec_bf": elec_bf}
        figures = parse_csv_figures(files, location, passthrough)
        return {"location": location, "figures": figures}
    except Exception as e:
        raise HTTPException(500, f"CSV parse error: {str(e)}")


@app.post("/generate")
async def generate_report(
    location: str = Form(...),
    figures:  str = Form(...),   # JSON string of (possibly edited) figures
    report_date: str = Form(...),
):
    """
    Step 2 — Receive (reviewed/edited) figures, calculate reconciliation, generate .docx.
    """
    if location not in PARSERS:
        raise HTTPException(400, f"Unknown location: {location}")
    try:
        figs = json.loads(figures)
    except Exception:
        raise HTTPException(400, "Invalid figures JSON")
    try:
        result = calculate(location, figs)
        buf = generate_xlsx(location, result, report_date)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f'attachment; filename="{location.upper()}_{report_date.replace("-","_")}_Debtors_Reconciliation.xlsx"'}
        )
    except Exception as e:
        raise HTTPException(500, f"Generation error: {str(e)}")
# ──────────────────────────────────────────────────────────────────────
# ADJUSTMENT TAB — independent of the reconciliation flow above.
# Diffs prior-month CLOSING debtors vs current-month OPENING debtors and
# returns Adjustment (1) plus the full invoice-level detail.
# ──────────────────────────────────────────────────────────────────────
@app.post("/adjustments")
async def adjustments(
    adjustment_month: str = Form(...),        # any date in the adjustment month, e.g. "2026-03-01"
    adj_close_csv: UploadFile = File(...),     # previous month CLOSING debtors export (.csv)
    adj_open_csv:  UploadFile = File(...),     # current month OPENING debtors export (.csv)
    location: str = Form(None),                # OPTIONAL — auto-detected from the CSV if omitted
):
    if location and location not in ISLAND_BY_LOCATION:
        raise HTTPException(400, f"Unknown location: {location}")
    try:
        month = datetime.fromisoformat(adjustment_month).date()
    except Exception:
        raise HTTPException(400, "adjustment_month must be ISO date, e.g. 2026-03-01")
    try:
        tmp_dir = tempfile.mkdtemp()
        close_path = os.path.join(tmp_dir, "close.csv")
        open_path  = os.path.join(tmp_dir, "open.csv")
        with open(close_path, "wb") as f:
            f.write(await adj_close_csv.read())
        with open(open_path, "wb") as f:
            f.write(await adj_open_csv.read())

        cutoff = datetime(month.year, month.month, 1)
        try:
            items, summary = identify(close_path, open_path, location=location, report_cutoff=cutoff)
        except ValueError as e:
            raise HTTPException(400, str(e))

        # Island name (auto-detected) drives the output filename when no location was picked.
        island_tag = (summary.get("island") or location or "ADJUSTMENT").upper().replace(" ", "_").replace("'", "")
        out_path = os.path.join(
            tmp_dir,
            f"{island_tag}_{month.strftime('%Y_%m')}_Adjustment_Details.xlsx"
        )
        write_xlsx(items, summary, location, month, out_path)
        import base64
        summary_b64 = base64.b64encode(json.dumps(summary).encode()).decode()
        return FileResponse(
            out_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(out_path),
            headers={"X-Adjustment-Total": str(summary["total_adjustment"]),
                     "X-Adjustment-Rows": str(summary["n_rows"]),
                     "X-Adjustment-Review": str(summary["n_review"]),
                     "X-Adjustment-Summary": summary_b64},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Adjustment error: {str(e)}")
# ──────────────────────────────────────────────────────────────────────
# ADJUSTMENT (2) TAB — the CLOSING-side plug.
# Adj(2) = Closing - Opening(after adj) - Sales - Credits + Collection.
# Attributes the plug to realised bill payments that reduced no open
# debtor balance (prior-period invoices absent from both debtor ledgers).
# ──────────────────────────────────────────────────────────────────────
@app.post("/adjustments2")
async def adjustments2(
    recon_month: str = Form(...),                # "2026-05"
    collection_csv:    UploadFile = File(...),   # collection transactions
    open_debtors_csv:  UploadFile = File(...),   # opening (prior-month closing) debtor detail
    close_debtors_csv: UploadFile = File(...),   # closing debtor detail
    sales_csv:         UploadFile = File(...),   # CURRENT-month sales (invoice-level)
    opening_after_adj: float = Form(None),       # control totals -> plug (all four enable it)
    closing: float = Form(None),
    sales:   float = Form(None),
    credits: float = Form(None),
):
    try:
        year, month = (int(p) for p in recon_month.split("-")[:2])
    except Exception:
        raise HTTPException(400, "recon_month must be YYYY-MM, e.g. 2026-05")
    try:
        coll       = load_collection_rows(await collection_csv.read())
        open_rows  = load_debtor_rows(await open_debtors_csv.read())
        close_rows = load_debtor_rows(await close_debtors_csv.read())
        sales_rows = load_debtor_rows(await sales_csv.read())

        plug = None
        if None not in (opening_after_adj, closing, sales, credits):
            plug = plug_from_totals(opening_after_adj, closing, sales, credits,
                                    realised_collection(coll))

        res = detect(coll, open_rows, close_rows, sales_rows, (year, month), plug=plug)

        return StreamingResponse(
            generate_xlsx_bytes(res),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "X-Adjustment2-Summary": summary_b64(res),
                "Content-Disposition": f'attachment; filename="Adjustment2_{recon_month}.xlsx"',
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Adjustment (2) error: {str(e)}")

// ── Config ────────────────────────────────────────────────────
const API_URL = "https://stelco-recon-api.onrender.com";

// ── State ─────────────────────────────────────────────────────
const state = {
  step:      1,
  location:  null,
  month:     null,
  date:      null,
  figures:   {},
  calcResult: null,
};

// ── File slots per location ────────────────────────────────────
const FILE_SLOTS = (function () {
  // Core CSV set — same for every location. MISC + prior-close are optional.
  const core = [
    { key: "open_csv",        label: "Opening Debtors (CSV)",  icon: "📂", desc: "Current-month opening debtor detail",           required: true  },
    { key: "close_csv",       label: "Closing Debtors (CSV)",  icon: "📁", desc: "Current-month closing debtor detail",           required: true  },
    { key: "sales_csv",       label: "Sales Report (CSV)",     icon: "📊", desc: "Current-month sales (invoice-level)",           required: true  },
    { key: "collection_csv",  label: "Collection (CSV)",       icon: "🧾", desc: "Collection transactions",                       required: true  },
    { key: "credits_csv",     label: "Credits (CSV)",          icon: "💳", desc: "Credits / Fine collection report",             required: false },
    { key: "prior_close_csv", label: "Prior Closing (CSV)",    icon: "🗂️", desc: "Prior-month closing — enables Adjustment (1)", required: false },
    { key: "misc_open_csv",   label: "MISC Opening (CSV)",     icon: "📂", desc: "MISC opening debtors",                          required: false },
    { key: "misc_close_csv",  label: "MISC Closing (CSV)",     icon: "📁", desc: "MISC closing debtors",                          required: false },
    { key: "misc_sales_csv",  label: "MISC Sales (CSV)",       icon: "📊", desc: "MISC sales",                                    required: false },
    { key: "misc_coll_csv",   label: "MISC Collection (CSV)",  icon: "🧾", desc: "MISC reconciled collection",                    required: false },
  ];
  const all = {};
  ["male","thilafushi","gulhi_falhu","other_islands","hulhumale"].forEach(loc => { all[loc] = core.slice(); });
  return all;
})();

const LOCATION_NAMES = {
  male:        "Male'",
  hulhumale:   "Hulhumale'",
  thilafushi:  "Thilafushi",
  gulhi_falhu: "Gulhi Falhu",
  other_islands: "Other Islands",
};

// Review fields shown in step 3
const REVIEW_FIELDS = [
  { key: "elec_bf",           label: "Balance b/f (prior month signed c/f — enter manually)",       misc_key: "misc_bf"           },
  { key: "elec_bfadj",        label: "Balance b/f (after adjustment)",                               misc_key: "misc_bfadj"        },
  { key: "elec_sales",        label: "Total Sales/Additional Revenue",                               misc_key: "misc_sales"        },
  { key: "elec_credits",      label: "Credits / Fine",                                               misc_key: "misc_credits"      },
  { key: "elec_collection",   label: "Collection for the month",                                     misc_key: "misc_collection"   },
  { key: "elec_close_system", label: "Debtors Balance c/f (from close.pdf — verify at month-end)",  misc_key: "misc_close_system" },
];

// Hulhumale extra fields
const HULHUMALE_EXTRA = [
  { key: "billing_system", label: "Billing System Collection", misc_key: null },
  { key: "blueridge",      label: "Blueridge Collections",     misc_key: null },
  { key: "wamco",          label: "WAMCO Collections",         misc_key: null },
];

// ── Helpers ────────────────────────────────────────────────────
const fmt = (n) => {
  if (n === null || n === undefined || n === "") return "";
  const num = parseFloat(n);
  if (isNaN(num)) return String(n);
  const abs = Math.abs(num);
  const s = abs.toLocaleString("en-US", { minimumFractionDigits: 2, maximumFractionDigits: 2 });
  return num < 0 ? `(${s})` : s;
};

const parseNum = (s) => {
  if (!s) return 0;
  return parseFloat(String(s).replace(/[^0-9.-]/g, "")) || 0;
};

// ── Reconciliation snapshot (live statement preview) ───────────
// Mirrors reconciliation/calculator.py row math for the right-side preview.
function computeRecon(f) {
  f = f || {};
  const g = k => parseNum(f[k]);
  const has = k => f[k] !== undefined && f[k] !== null && f[k] !== "";
  const e_bf = g("elec_bf"), m_bf = g("misc_bf");
  const e_bfadj = has("elec_bfadj") ? g("elec_bfadj") : e_bf;
  const m_bfadj = has("misc_bfadj") ? g("misc_bfadj") : m_bf;
  const e_sales = g("elec_sales"), m_sales = g("misc_sales");
  const e_cr = g("elec_credits"), m_cr = g("misc_credits");
  const e_disc = g("elec_discount"), m_disc = g("misc_discount");
  const e_coll = g("elec_collection"), m_coll = g("misc_collection");
  const e_cf = g("elec_close_system"), m_cf = g("misc_close_system");
  const e_sub1 = e_bfadj + e_sales, m_sub1 = m_bfadj + m_sales;
  const e_sub2 = e_sub1 + e_cr + e_disc, m_sub2 = m_sub1 + m_cr + m_disc;
  const e_adj2 = e_cf - (e_sub2 - e_coll), m_adj2 = m_cf - (m_sub2 - m_coll);
  return [
    { label: "Balance b/f",            e: e_bf,    m: m_bf },
    { label: "Adjustments (1)",        e: e_bfadj - e_bf, m: m_bfadj - m_bf },
    { label: "Balance b/f (adj.)",     e: e_bfadj, m: m_bfadj, bold: true },
    { label: "Total Sales",            e: e_sales, m: m_sales },
    { label: "",                       e: e_sub1,  m: m_sub1, sub: true },
    { label: "Credits / Fine",         e: e_cr,    m: m_cr },
    { label: "Discount",               e: e_disc,  m: m_disc },
    { label: "",                       e: e_sub2,  m: m_sub2, sub: true },
    { label: "Collection",             e: -e_coll, m: -m_coll },
    { label: "Adjustments (2)",        e: e_adj2,  m: m_adj2 },
    { label: "Debtors Balance c/f",    e: e_cf,    m: m_cf, final: true },
  ];
}

function renderSnapshot() {
  const el = document.getElementById("reconSnapshot");
  if (!el) return;
  const hasFigures = state.figures && Object.keys(state.figures).length > 0;
  const rows = computeRecon(state.figures);
  const monthStr = state.month
    ? new Date(state.month + "-01").toLocaleString("en-US", { month: "long", year: "numeric" })
    : "—";
  const loc = LOCATION_NAMES[state.location] || "—";
  let html = `<div class="snap-head"><span>${loc}</span><span>${monthStr}</span></div>`;
  html += `<table class="snap-table"><thead><tr><th></th><th>Elec.</th><th>MISC</th></tr></thead><tbody>`;
  for (const r of rows) {
    const cls = `${r.bold ? "snap-bold " : ""}${r.sub ? "snap-sub " : ""}${r.final ? "snap-final " : ""}`.trim();
    const ev = hasFigures ? fmt(r.e) : "—";
    const mv = hasFigures ? fmt(r.m) : "—";
    html += `<tr class="${cls}"><td>${r.label}</td><td>${ev}</td><td>${mv}</td></tr>`;
  }
  html += `</tbody></table>`;
  if (!hasFigures) html += `<div class="snap-hint">Preview — fills in after Parse Files</div>`;
  el.innerHTML = html;
}

// ── Step navigation ────────────────────────────────────────────
function goToStep(n) {
  state.step = n;
  document.querySelectorAll(".step-panel").forEach(p => p.classList.remove("active"));
  document.getElementById(`panel${n}`).classList.add("active");

  document.querySelectorAll(".step-item").forEach(item => {
    const s = parseInt(item.dataset.step);
    item.classList.remove("active", "done");
    if (s === n) item.classList.add("active");
    else if (s < n) item.classList.add("done");
  });

  window.scrollTo(0, 0);

  if (n === 3) startParsing();
  if (n === 4) { renderSummary(); renderSnapshot(); }
}

// ── STEP 1 ─────────────────────────────────────────────────────
document.querySelectorAll(".loc-card").forEach(card => {
  card.addEventListener("click", () => {
    document.querySelectorAll(".loc-card").forEach(c => c.classList.remove("selected"));
    card.classList.add("selected");
    state.location = card.dataset.location;
    checkStep1();
  });
});

document.getElementById("reportMonth").addEventListener("change", (e) => {
  state.month = e.target.value;
  checkStep1();
});
document.getElementById("reportDate").addEventListener("change", (e) => {
  state.date = e.target.value;
  checkStep1();
});

function checkStep1() {
  document.getElementById("step1Next").disabled = !(state.location && state.month && state.date);
}

document.getElementById("step1Next").addEventListener("click", () => {
  renderUploadSlots();
  document.getElementById("locationLabel").textContent = LOCATION_NAMES[state.location];
  goToStep(2);
});

// ── STEP 2 ─────────────────────────────────────────────────────
const uploadedFiles = {};

function renderUploadSlots() {
  const grid = document.getElementById("uploadGrid");
  grid.innerHTML = "";
  const slots = FILE_SLOTS[state.location] || [];

  slots.forEach(slot => {
    const div = document.createElement("div");
    div.className = "upload-slot" + (slot.required ? "" : " optional");
    div.id = `slot_${slot.key}`;
    div.innerHTML = `
      <div class="slot-icon">${slot.icon}</div>
      <div class="slot-info">
        <div class="slot-name">${slot.label}</div>
        <div class="slot-desc">${slot.desc}</div>
      </div>
      <span class="slot-badge ${slot.required ? "req" : "opt"}">${slot.required ? "Required" : "Optional"}</span>
      <div class="file-input-wrap">
        <label class="file-btn" id="btn_${slot.key}">
          Choose
          <input type="file" accept=".csv" data-key="${slot.key}" onchange="onFileChosen(this)">
        </label>
      </div>`;
    grid.appendChild(div);
  });
}

function onFileChosen(input) {
  const key = input.dataset.key;
  const file = input.files[0];
  if (!file) return;
  uploadedFiles[key] = file;

  const slot = document.getElementById(`slot_${key}`);
  slot.classList.add("filled");
  const btn = document.getElementById(`btn_${key}`);
  btn.classList.add("chosen");
  btn.childNodes[0].textContent = file.name.length > 16 ? file.name.slice(0, 14) + "…" : file.name;

  checkStep2();
}

function checkStep2() {
  const slots = FILE_SLOTS[state.location] || [];
  const required = slots.filter(s => s.required).map(s => s.key);
  const allFilled = required.every(k => uploadedFiles[k]);
  document.getElementById("step2Next").disabled = !allFilled;
}

document.getElementById("step2Next").addEventListener("click", () => goToStep(3));

// ── STEP 3: Parse + Review ─────────────────────────────────────
async function startParsing() {
  document.getElementById("parsingStatus").style.display = "flex";
  document.getElementById("reviewWrap").style.display = "none";
  document.getElementById("step3Next").style.display = "none";

  // Step 1: Wake up the server first
  document.querySelector("#parsingStatus span").textContent = "Waking up server (may take ~60 seconds on first use)…";
  try { await fetch(`${API_URL}/`, { signal: AbortSignal.timeout(90000) }); } catch(e) {}

  // Step 2: Send the files for parsing
  document.querySelector("#parsingStatus span").textContent = "Parsing PDFs…";

  try {
    const form = new FormData();
    form.append("location", state.location);
    form.append("report_date", state.date);
    // pass through typed values so the backend can fold them in
    ["blueridge","wamco","elec_bf"].forEach(k => {
      const v = state.figures[k];
      if (v !== undefined && v !== null && v !== "") form.append(k, v);
    });

    const slots = FILE_SLOTS[state.location] || [];
    for (const slot of slots) {
      if (uploadedFiles[slot.key]) {
        form.append(slot.key, uploadedFiles[slot.key]);
      }
    }

    const res = await fetch(`${API_URL}/parse_csv`, {
      method: "POST",
      body: form,
      signal: AbortSignal.timeout(120000)  // 2 minute timeout
    });
    if (!res.ok) throw new Error(await res.text());
    const data = await res.json();
    state.figures = data.figures || {};
    // Locations whose parser doesn't split b/f vs b/f-after-adjustment:
    // default bfadj to bf so Adjustments(1) is 0 unless the user edits.
    if (state.figures.elec_bfadj == null) state.figures.elec_bfadj = state.figures.elec_bf ?? 0;
    if (state.figures.misc_bfadj == null) state.figures.misc_bfadj = state.figures.misc_bf ?? 0;

  } catch (err) {
    // If backend unreachable, use empty figures so user can still enter manually
    console.warn("Parse failed, using empty figures:", err.message);
    state.figures = {};
  }

  renderReviewTable();
  renderSnapshot();
  document.getElementById("parsingStatus").style.display = "none";
  document.getElementById("reviewWrap").style.display = "block";
  document.getElementById("step3Next").style.display = "inline-block";
}

function renderReviewTable() {
  const tbody = document.getElementById("reviewBody");
  tbody.innerHTML = "";

  let fields = [...REVIEW_FIELDS];
  if (state.location === "hulhumale") fields = [...HULHUMALE_EXTRA, ...fields];

  fields.forEach(field => {
    const elecVal = state.figures[field.key] ?? null;
    const miscVal = field.misc_key ? (state.figures[field.misc_key] ?? null) : null;
    const elecMissing = elecVal === null || elecVal === 0;

    const tr = document.createElement("tr");
    tr.innerHTML = `
      <td><strong>${field.label}</strong></td>
      <td>
        <input class="editable-field ${elecMissing ? "needs-input" : ""}"
               data-key="${field.key}"
               value="${elecVal !== null && elecVal !== 0 ? elecVal : ""}"
               placeholder="${elecMissing ? "Enter value" : ""}">
      </td>
      <td>
        ${field.misc_key
          ? `<input class="editable-field" data-key="${field.misc_key}" value="${miscVal !== null ? miscVal : 0}">`
          : `<span style="color:var(--muted)">—</span>`
        }
      </td>`;
    tbody.appendChild(tr);
  });

  // Live-update state on edit
  tbody.querySelectorAll(".editable-field").forEach(input => {
    input.addEventListener("change", () => {
      state.figures[input.dataset.key] = parseNum(input.value);
      input.classList.remove("needs-input");
      renderSnapshot();
    });
  });
}

// ── STEP 4: Summary + Generate ─────────────────────────────────
function renderSummary() {
  const monthStr = state.month
    ? new Date(state.month + "-01").toLocaleString("en-US", { month: "long", year: "numeric" })
    : "—";

  document.getElementById("summaryLocation").textContent = LOCATION_NAMES[state.location] || "—";
  document.getElementById("summaryMonth").textContent = monthStr;

  const grid = document.getElementById("summaryGrid");
  grid.innerHTML = "";

  const highlights = [
    { label: "ELECTRICITY Balance b/f",  val: state.figures["elec_bf"] },
    { label: "MISC Balance b/f",         val: state.figures["misc_bf"] },
    { label: "Electricity Sales",        val: state.figures["elec_sales"] },
    { label: "MISC Sales",               val: state.figures["misc_sales"] },
    { label: "Electricity Collection",   val: state.figures["elec_collection"] },
    { label: "System Closing (Elec)",    val: state.figures["elec_close_system"], highlight: true },
    { label: "System Closing (MISC)",    val: state.figures["misc_close_system"], highlight: true },
  ];

  highlights.forEach(item => {
    if (item.val === undefined || item.val === null) return;
    const div = document.createElement("div");
    div.className = "summary-item";
    div.innerHTML = `
      <div class="summary-item-label">${item.label}</div>
      <div class="summary-item-val ${item.highlight ? "highlight" : ""}">MRF ${fmt(item.val)}</div>`;
    grid.appendChild(div);
  });
}

async function generateReport() {
  const btn  = document.getElementById("generateBtn");
  const text = document.getElementById("generateBtnText");
  const successMsg = document.getElementById("successMsg");
  const errorMsg   = document.getElementById("errorMsg");

  btn.disabled = true;
  text.textContent = "Generating…";
  successMsg.style.display = "none";
  errorMsg.style.display   = "none";

  // Read latest values from review table inputs
  document.querySelectorAll("#reviewBody .editable-field").forEach(input => {
    state.figures[input.dataset.key] = parseNum(input.value);
  });

  const form = new FormData();
  form.append("location",    state.location);
  form.append("figures",     JSON.stringify(state.figures));
  form.append("report_date", state.date);

  try {
    const res = await fetch(`${API_URL}/generate`, {
      method: "POST",
      body: form,
      signal: AbortSignal.timeout(120000)
    });
    if (!res.ok) throw new Error(await res.text());

    const blob = await res.blob();
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement("a");
    const monthStr = state.month ? state.month.replace("-", "_") : "report";
    a.href     = url;
    a.download = `${state.location.toUpperCase()}_${monthStr}_Debtors_Reconciliation.xlsx`;
    a.click();
    URL.revokeObjectURL(url);

    successMsg.style.display = "block";
  } catch (err) {
    errorMsg.textContent = `Error: ${err.message}`;
    errorMsg.style.display = "block";
  } finally {
    btn.disabled = false;
    text.textContent = "⬇ Download .xlsx Report";
  }
}

"""
reconciliation/csv_parser.py
Replaces the PDF parsers: computes the same `figures` dict the web app's
review step already expects, but straight from CSV exports.
"""
import csv, io


def _text(src):
    if hasattr(src, "read"):
        d = src.read()
        return d.decode("utf-8-sig", "replace") if isinstance(d, (bytes, bytearray)) else d
    if isinstance(src, (bytes, bytearray)):
        return bytes(src).decode("utf-8-sig", "replace")
    with open(src, encoding="utf-8-sig", errors="replace", newline="") as fh:
        return fh.read()


def _rows(src):
    if src is None:
        return []
    return list(csv.DictReader(io.StringIO(_text(src))))


def _num(x):
    x = (x or "").strip().replace(",", "")
    try:
        return float(x) if x else 0.0
    except ValueError:
        return 0.0


def _find(fields, preferred, keys):
    fields = [f for f in (fields or []) if f]
    for f in fields:
        if f.strip().lstrip("\ufeff").upper() == preferred.upper():
            return f
    low = {f.strip().lstrip("\ufeff").lower(): f for f in fields}
    if preferred.lower() in low:
        return low[preferred.lower()]
    for f in fields:
        if any(k in f.strip().lower() for k in keys):
            return f
    return None


def _sum(rows, preferred, keys):
    if not rows:
        return 0.0
    col = _find(rows[0].keys(), preferred, keys)
    if not col:
        return 0.0
    return round(sum(_num(r.get(col)) for r in rows), 2)


ELEC_DESCRIPTIONS = {"ELECTRICITY SALE", "FINE", "CHARGES FROM PREV BILL"}


def _realised_collection(rows):
    if not rows:
        return 0.0
    hdr = rows[0].keys()
    c_ord = _find(hdr, "ORD", ["ord"])
    c_canc = _find(hdr, "CANCEL_DATE", ["cancel"])
    c_desc = _find(hdr, "DESCRIPTION", ["description", "cdt_desc"])
    c_amt = _find(hdr, "COLLECT_AMOUNT", ["collect_amount", "collect"])
    c_comp = _find(hdr, "AMOUNT", ["amount"])
    total = 0.0
    for r in rows:
        if c_canc and (r.get(c_canc) or "").strip():
            continue
        if c_ord:
            if (r.get(c_ord) or "").strip() in ("1", "2"):
                total += _num(r.get(c_amt))
        elif c_desc:
            if (r.get(c_desc) or "").strip().upper() in ELEC_DESCRIPTIONS:
                total += _num(r.get(c_comp) if c_comp else r.get(c_amt))
        else:
            total += _num(r.get(c_amt))
    return round(total, 2)


def parse_csv_figures(files, location, passthrough=None):
    passthrough = passthrough or {}
    open_rows = _rows(files.get("open_csv"))
    close_rows = _rows(files.get("close_csv"))
    sales_rows = _rows(files.get("sales_csv"))
    coll_rows = _rows(files.get("collection_csv"))
    credits_rows = _rows(files.get("credits_csv"))
    prior_rows = _rows(files.get("prior_close_csv"))

    elec_bfadj = _sum(open_rows, "BALANCE_AMT", ["balance", "outstand"])
    elec_close = _sum(close_rows, "BALANCE_AMT", ["balance", "outstand"])
    elec_sales = _sum(sales_rows, "AMOUNT", ["amount"])
    elec_credits = _sum(credits_rows, "AMOUNT", ["amount"])
    elec_billing = _realised_collection(coll_rows)

    if prior_rows:
        elec_bf = _sum(prior_rows, "BALANCE_AMT", ["balance", "outstand"])
    else:
        elec_bf = _num(passthrough.get("elec_bf")) or elec_bfadj

    blueridge = _num(passthrough.get("blueridge"))
    wamco = _num(passthrough.get("wamco"))
    if location == "hulhumale":
        elec_collection = round(elec_billing + blueridge + wamco, 2)
    else:
        elec_collection = elec_billing

    figures = {
        "elec_bf": elec_bf,
        "elec_bfadj": elec_bfadj,
        "elec_sales": elec_sales,
        "elec_credits": elec_credits,
        "elec_discount": _num(passthrough.get("elec_discount")),
        "elec_collection": elec_collection,
        "elec_close_system": elec_close,
    }

    misc_open = _rows(files.get("misc_open_csv"))
    misc_close = _rows(files.get("misc_close_csv"))
    misc_sales = _rows(files.get("misc_sales_csv"))
    misc_coll = _rows(files.get("misc_coll_csv"))
    if misc_open or misc_close:
        m_bfadj = _sum(misc_open, "BALANCE_AMT", ["balance", "outstand"])
        figures.update({
            "misc_bf": m_bfadj,
            "misc_bfadj": m_bfadj,
            "misc_sales": _sum(misc_sales, "ITM_AMOUNT", ["itm_amount", "amount"]),
            "misc_credits": _num(passthrough.get("misc_credits")),
            "misc_discount": _num(passthrough.get("misc_discount")),
            "misc_collection": _sum(misc_coll, "AMOUNT", ["amount"]),
            "misc_close_system": _sum(misc_close, "BALANCE_AMT", ["balance", "outstand"]),
        })
    else:
        for k in ("misc_bf", "misc_bfadj", "misc_sales", "misc_credits",
                  "misc_discount", "misc_collection", "misc_close_system"):
            figures[k] = _num(passthrough.get(k))

    if location == "hulhumale":
        figures["billing_system"] = elec_billing
        figures["blueridge"] = blueridge
        figures["wamco"] = wamco
    return figures

"""
reconciliation/generator_xlsx.py
Excel version of the Debtors Reconciliation Statement.
Consumes calculate(location, figures) -> result (same dict as before) and
produces an .xlsx laid out like the Word statement (electricity + MISC columns).
"""
import io


def _fmt_num(n):
    return round(float(n or 0), 2)


def generate_xlsx(location, result, report_date):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    bold = Font(name="Arial", size=10, bold=True)
    normal = Font(name="Arial", size=10)
    red = Font(name="Arial", size=10, color="CC0000")
    red_bold = Font(name="Arial", size=10, bold=True, color="CC0000")
    grey = PatternFill("solid", fgColor="D9D9D9")
    light = PatternFill("solid", fgColor="EEEEEE")
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")
    thin_top = Border(top=Side(style="thin"))
    dbl_bottom = Border(bottom=Side(style="double"))

    # Header block
    ws["A1"] = "Customer Services and Billing Department"; ws["A1"].font = bold
    ws["A2"] = "    STELCO"; ws["A2"].font = normal
    ws["A4"] = f"Date: {report_date}"; ws["A4"].font = normal
    ws["A6"] = f"Debtors Reconciliation Statement - {result['location_name']}"
    ws["A6"].font = Font(name="Arial", size=11, bold=True, underline="single")
    ws["A7"] = report_date; ws["A7"].font = bold

    # Column headers
    hr = 9
    ws.cell(hr, 2, "ELECTRICITY (MRF)").font = bold
    ws.cell(hr, 3, "MISC. (MRF)").font = bold
    ws.cell(hr, 2).alignment = center
    ws.cell(hr, 3).alignment = center
    ws.cell(hr, 2).fill = grey
    ws.cell(hr, 3).fill = grey
    ws.cell(hr, 1).fill = grey

    r = hr + 1
    for row in result["rows"]:
        is_sub = row.get("subtotal", False)
        is_final = row.get("final", False)
        is_bold = row.get("bold", False) or is_sub or is_final
        ws.cell(r, 1, row["label"]).font = bold if is_bold else normal
        for ci, key in ((2, "elec"), (3, "misc")):
            v = _fmt_num(row[key])
            cell = ws.cell(r, ci, v)
            cell.number_format = "#,##0.00;(#,##0.00)"
            cell.alignment = right
            if v < 0:
                cell.font = red_bold if is_bold else red
            else:
                cell.font = bold if is_bold else normal
            if is_sub:
                cell.fill = light; cell.border = thin_top
            if is_final:
                cell.fill = grey; cell.border = dbl_bottom
        if is_sub:
            ws.cell(r, 1).fill = light
        if is_final:
            ws.cell(r, 1).fill = grey
        r += 1

    # footnote
    r += 1
    ws.cell(r, 1, "*Credit invoice in the Total Sales").font = Font(name="Arial", size=8, italic=True, color="555555")

    # signatures
    r += 3
    sigs = [("Prepared By:", "Hamza Abdul Sattar", "Admin. Supervisor"),
            ("Checked By:", "Ali Amir", "Deputy Service Manager"),
            ("Approved By:", "Hussain Waheed", "General Manager")]
    for ci, (lab, name, title) in enumerate(sigs, start=1):
        ws.cell(r, ci, lab).font = bold
        ws.cell(r + 4, ci, name).font = bold
        ws.cell(r + 5, ci, title).font = normal

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
